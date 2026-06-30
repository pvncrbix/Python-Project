from openpyxl import load_workbook
from tkinter import *

wb = load_workbook('C:\\Users\\Admin\\Desktop\\excel.xlsx')
ws = wb.active

# Set header row
def init_excel():
    headers = ["Name", "Course", "Semester", "Form No.", "Contact No.", "Email", "Address"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i).value = h
    wb.save('C:\\Users\\Admin\\Desktop\\excel.xlsx')

# Add form data
def insert_data():
    if all(f.get() for f in entries):
        row = ws.max_row + 1
        for i, f in enumerate(entries, 1):
            ws.cell(row=row, column=i).value = f.get()
        wb.save('C:\\Users\\Admin\\Desktop\\excel.xlsx')
        clear_fields()
    else:
        print("Please fill all fields.")

# Clear form
def clear_fields():
    for f in entries:
        f.delete(0, END)

# Move to next field
def focus_next(entry):
    entry.focus_set()

init_excel()
root = Tk()
root.title("Registration Form")
root.geometry("500x300")
root.config(bg="light green")

labels = ["Name", "Course", "Semester", "Form No.", "Contact No.", "Email", "Address"]
entries = [Entry(root) for _ in labels]

for i, lbl in enumerate(labels):
    Label(root, text=lbl, bg="light green").grid(row=i+1, column=0)
    entries[i].grid(row=i+1, column=1, ipadx=100)
    if i < len(labels) - 1:
        entries[i].bind("<Return>", lambda e, nf=entries[i+1]: focus_next(nf))

Button(root, text="Submit", fg="black", bg="red", command=insert_data).grid(row=len(labels)+1, column=1)

root.mainloop()